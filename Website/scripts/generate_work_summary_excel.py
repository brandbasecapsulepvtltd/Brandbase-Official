"""Generate merged June–July 2026 day-wise work summary Excel (IST) — one row per day."""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

OUTPUT = r"c:\BrandBase\Brandbase Capsule Handover\Brandbase\Website\BrandBase-Daywise-Work-Summary-June-July-2026.xlsx"

# One concise summary per working day (IST)
DAILY_ROWS = [
    ("2026-06-27", "Saturday", "Client + Backend + Docs",
     "Started full public-site enhancement pass: home SEO/JSON-LD, site-wide foundation (#FF6600, skip link, canonical fix), about/contact/appointment/services/blog/footer UI work, backend scheduler & CORS; FIXES-HANDOVER.md created."),
    ("2026-06-28", "Sunday", "Client + Backend",
     "Major SEO & UX pass across 20+ pages: dynamic sitemap, event calendar + detail, expo/Mumbai landing, legal pages, portfolio listing/detail, services/AV/blogs metadata, imageUtils + service images, hydration fixes on service pages."),
    ("2026-06-29", "Monday", "Client + Docs",
     "Global search fix (no gibberish matches); work summary & handover docs updated."),
    ("2026-07-01", "Wednesday", "Client",
     "Minor client maintenance / small file updates."),
    ("2026-07-03", "Friday", "Client + Docs",
     "Pre-production checklist; hid phone & address on contact (form-only); home hero slides redesigned with services-style animations."),
    ("2026-07-06", "Monday", "Client",
     "Minor client updates / prep work."),
    ("2026-07-07", "Tuesday", "Client + Backend + Docs",
     "HGH India 2026 blog + portfolio + home recent work; SEO landing pages (5 services, Mumbai + India); exhibition-first service order; home slide images (web/event/AV); console fixes; event-management portfolio image; portfolio overlap bug investigation."),
    ("2026-07-08", "Wednesday", "Client",
     "Portfolio page rebuilt (v3 grid); fixed OTHER/EVENT MANAGEMENT overlapping card; cropped event-management thumbnail; footer/hydration & dev-server fixes."),
    ("2026-07-09", "Thursday", "Docs",
     "Deploy scope documented (client + backend); backend deploy guide; merged June–July work summary Excel."),
]

HEADERS = ["Date", "Day", "Apps", "Work Summary"]


def style_header(ws, cols, fill_color="1F4E79"):
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    for c in range(1, cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def auto_width(ws, widths=None):
    if widths:
        for letter, w in widths.items():
            ws.column_dimensions[letter].width = w
        return
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 55)


def main():
    wb = Workbook()

    # Overview
    ws0 = wb.active
    ws0.title = "Overview"
    ws0["A1"] = "BrandBase Capsule — Work Summary"
    ws0["A1"].font = Font(bold=True, size=16, color="1F4E79")
    ws0["A2"] = "Period: June – July 2026 (IST) · One row per working day"
    ws0["A3"] = f"Generated: {datetime.now().strftime('%d %B %Y, %I:%M %p IST')}"
    ws0["A5"] = "Deploy"
    ws0["B5"] = "1) Client  2) Backend  3) Skip Admin"
    ws0["A6"] = "Docs"
    ws0["B6"] = "FIXES-HANDOVER.md · WORK-SUMMARY-SAT-AND-TODAY.md"
    ws0.column_dimensions["A"].width = 14
    ws0.column_dimensions["B"].width = 60
    for r in (5, 6):
        ws0.cell(r, 1).font = Font(bold=True)

    # Day-wise (one row per day)
    ws = wb.create_sheet("Day-wise (IST)")
    ws.append(HEADERS)
    style_header(ws, len(HEADERS))
    for row in DAILY_ROWS:
        ws.append(list(row))
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(HEADERS) + 1):
            ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    auto_width(ws, {"A": 14, "B": 14, "C": 22, "D": 90})

    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")
    print(f"Days: {len(DAILY_ROWS)}")


if __name__ == "__main__":
    main()
